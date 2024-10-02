
import openpyxl

# TODO 1 - Open current excel file and create new file for output. Create two list for each column

input_ssh = openpyxl.load_workbook('sample_data.xlsx')
input_sheet = input_ssh.active
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
col_1 = []
col_2 = []

# TODO 2 - Create nested loop to loop through rows and columns.
#  Insert column 1 values to col_1 and column 2 values to col_2

for row in range(1, input_sheet.max_row + 1):
    for col in range(1, input_sheet.max_column + 1):
        if col == 1:
            col_val_1 = input_sheet.cell(row=row, column=col).value
            col_1.append(col_val_1)
        else:
            col_val_2 = input_sheet.cell(row=row, column=col).value
            col_2.append(col_val_2)


# TODO 3 - To transpose we need to loop through each list and insert values to rows of new excel file and save output
col_count = 1
col_count_2 = 1
for item in col_1:
    output_sheet.cell(row=1, column=col_count).value = item
    col_count += 1
for item in col_2:
    output_sheet.cell(row=2, column=col_count_2).value = item
    col_count_2 += 1
output_wb.save('output.xlsx')