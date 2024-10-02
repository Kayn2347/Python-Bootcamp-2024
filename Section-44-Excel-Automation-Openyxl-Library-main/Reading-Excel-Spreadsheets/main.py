from openpyxl import load_workbook


workbook = load_workbook("AmazonReviews.xlsx")
sheet = workbook.active
cell1 = sheet.cell(row=1, column=1)
print(cell1.value)