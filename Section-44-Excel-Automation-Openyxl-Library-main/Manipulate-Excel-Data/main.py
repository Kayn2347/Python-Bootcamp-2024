from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pprint


workbook = load_workbook("AmazonReviews.xlsx")
sheet = workbook.active
# for cols in cols:
#     for col in cols:
#         print(col.value)


for value in sheet.iter_rows(min_row=1, max_row=1, value_only=True):
  print(value)

review = {}
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=8, values_only=True):
    review_id = row[0]
    review = {
        "profile_name": row[1],
        "title": row[4],
        "rating": row[5]
    }
    review[review_id] = review

pprint.pprint(review)