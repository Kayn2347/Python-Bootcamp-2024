from openpyxl import load_workbookd
from openpyxl.utils import get_column_letter, column_index_from_string


workbook = load_workbook("AmazonReviews.xlsx")
sheet = workbook.active
# for cols in cols:
#     for col in cols:
#         print(col.value)


print(column_index_from_string("AA"))