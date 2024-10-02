from openpyxl import load_workbook


workbook = load_workbook("AmazonReviews.xlsx")
sheet = workbook.active
for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10,values_only= True):
print(row)