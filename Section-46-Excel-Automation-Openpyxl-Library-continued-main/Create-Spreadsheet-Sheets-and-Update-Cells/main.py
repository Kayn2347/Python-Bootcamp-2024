from openpyxl import Workbook


new_workbook = Workbook()
sheet = new_workbook.active
sheet["A1"] = "name"
sheet["A2"] = "Kayn"
sheet["B1"] = "surname"
sheet["B2"] = "Abo-Abo"

sheet["B11"] = "testing"

new_workbook.create_sheet("New Sheet")
print(new_workbook.sheetnames)
new_sheet = new_workbook["New Sheet"]
new_sheet.title = "Renamed new sheet"
new_sheet["A1"] = "id"
new_sheet["B1"] = "address"
new_sheet["C14"] = "new sheet testing"

new_workbook.remove(sheet)
new_workbook.copy_worksheet(new_sheet)
print(new_workbook.sheetnames)
new_workbook.save("new_spreadsheet.xlsx")