# TODO 1 - Open and read the cells of an Excel document with openpyxl module
from openpyxl import load_workbook



print("Opening Excel spreadsheet...")
spreadsheet = load_workbook("transactions.xlsx")
sheet = spreadsheet.active
supplier_data = {}
print("Reading rows from the Excel spreadsheet..")
for row_num in range(2, sheet.max_row+1):
    transaction_type = sheet["B"+str(row_num)].value
    supplier_name = sheet["C"+str(row_num)].value
    amount = sheet["D"+str(row_num)].value